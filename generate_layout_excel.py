"""
Generates: Malayalam_Radial_Keyboard_Layout.xlsx
Sheets:
  1. Layout Overview  — full tree: root segment → position → each character
  2. Root Ring        — 12 root segments with angles and type
  3. Vowels (സ്വ)
  4. ക-വർഗ്ഗം
  5. ച-വർഗ്ഗം
  6. ട-വർഗ്ഗം
  7. ത-വർഗ്ഗം
  8. പ-വർഗ്ഗം
  9. അവർഗ്ഗ്യം (യ)
  10. മലയാള-വിശേഷ (ള)
  11. Malayalam Numerals (൦)
  12. Instructions
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_TEXT

# ── Palette ───────────────────────────────────────────────────────────────────
TEAL_DARK   = "005F56"
TEAL_MID    = "00897B"
TEAL_LIGHT  = "B2DFDB"
GREEN_HI    = "4CAF50"
GREEN_LIGHT = "C8E6C9"
CHAR_BG     = "E8F5E9"
HEADER_FG   = "FFFFFF"
BORDER_CLR  = "90A4AE"
WHITE       = "FFFFFF"
GREY_LIGHT  = "F5F5F5"
AMBER       = "FF8F00"
AMBER_LIGHT = "FFF8E1"
RED_LIGHT   = "FFEBEE"
BLUE_LIGHT  = "E3F2FD"
PURPLE_LIGHT= "F3E5F5"
ORANGE_LIGHT= "FFF3E0"
CYAN_LIGHT  = "E0F7FA"
LIME_LIGHT  = "F9FBE7"
PINK_LIGHT  = "FCE4EC"
YELLOW_LIGHT= "FFFDE7"

# Per-family background colors
FAMILY_COLORS = {
    "സ്വ":  BLUE_LIGHT,
    "ക":    GREEN_LIGHT,
    "ച":    AMBER_LIGHT,
    "ട":    ORANGE_LIGHT,
    "ത":    CYAN_LIGHT,
    "പ":    PURPLE_LIGHT,
    "യ":    LIME_LIGHT,
    "ള":    PINK_LIGHT,
    "ം":    GREY_LIGHT,
    "⌫":    RED_LIGHT,
    "SPC":  GREY_LIGHT,
    "൦":    YELLOW_LIGHT,
}

# ── Layout data (mirrors KeyboardLayout.kt exactly) ───────────────────────────
ROOT_SEGMENTS = [
    {
        "root_index": 1,
        "label":      "സ്വ",
        "group_ml":   "സ്വരങ്ങൾ",
        "group_en":   "Vowels",
        "type":       "Drillable",
        "start_angle": -90,
        "sweep_angle":  30,
        "children":   ["അ","ആ","ഇ","ഈ","ഉ","ഊ","ഋ","എ","ഏ","ഐ","ഒ","ഓ","ഔ"],
        "phonetic":   ["a","aa","i","ee","u","oo","ri","e","ey","ai","o","oo","au"],
    },
    {
        "root_index": 2,
        "label":      "ക",
        "group_ml":   "ക-വർഗ്ഗം",
        "group_en":   "Velar (ka-vargam)",
        "type":       "Drillable",
        "start_angle": -60,
        "sweep_angle":  30,
        "children":   ["ക","ഖ","ഗ","ഘ","ങ"],
        "phonetic":   ["ka","kha","ga","gha","nga"],
    },
    {
        "root_index": 3,
        "label":      "ച",
        "group_ml":   "ച-വർഗ്ഗം",
        "group_en":   "Palatal (cha-vargam)",
        "type":       "Drillable",
        "start_angle": -30,
        "sweep_angle":  30,
        "children":   ["ച","ഛ","ജ","ഝ","ഞ"],
        "phonetic":   ["cha","chha","ja","jha","nya"],
    },
    {
        "root_index": 4,
        "label":      "ട",
        "group_ml":   "ട-വർഗ്ഗം",
        "group_en":   "Retroflex (ta-vargam)",
        "type":       "Drillable",
        "start_angle":  0,
        "sweep_angle":  30,
        "children":   ["ട","ഠ","ഡ","ഢ","ണ"],
        "phonetic":   ["tta","ttha","dda","ddha","nna"],
    },
    {
        "root_index": 5,
        "label":      "ത",
        "group_ml":   "ത-വർഗ്ഗം",
        "group_en":   "Dental (tha-vargam)",
        "type":       "Drillable",
        "start_angle":  30,
        "sweep_angle":  30,
        "children":   ["ത","ഥ","ദ","ധ","ന"],
        "phonetic":   ["tha","thha","da","dha","na"],
    },
    {
        "root_index": 6,
        "label":      "പ",
        "group_ml":   "പ-വർഗ്ഗം",
        "group_en":   "Labial (pa-vargam)",
        "type":       "Drillable",
        "start_angle":  60,
        "sweep_angle":  30,
        "children":   ["പ","ഫ","ബ","ഭ","മ"],
        "phonetic":   ["pa","pha","ba","bha","ma"],
    },
    {
        "root_index": 7,
        "label":      "യ",
        "group_ml":   "അവർഗ്ഗ്യം",
        "group_en":   "Semi-vowels & Sibilants",
        "type":       "Drillable",
        "start_angle":  90,
        "sweep_angle":  30,
        "children":   ["യ","ര","ല","വ","ശ","ഷ","സ","ഹ"],
        "phonetic":   ["ya","ra","la","va","sha","ssa","sa","ha"],
    },
    {
        "root_index": 8,
        "label":      "ള",
        "group_ml":   "മലയാള-വിശേഷ",
        "group_en":   "Malayalam-Specific + Chillus",
        "type":       "Drillable",
        "start_angle": 120,
        "sweep_angle":  30,
        "children":   ["ള","ഴ","റ","ൺ","ൻ","ർ","ൽ","ൾ"],
        "phonetic":   ["lla","zha","rra","chillu-nn","chillu-n","chillu-r","chillu-l","chillu-ll"],
    },
    {
        "root_index": 9,
        "label":      "ം",
        "group_ml":   "അനുസ്വാരം",
        "group_en":   "Anusvara (leaf)",
        "type":       "Leaf",
        "start_angle": 150,
        "sweep_angle":  30,
        "children":   [],
        "phonetic":   ["m̐ (nasal)"],
    },
    {
        "root_index": 10,
        "label":      "⌫",
        "group_ml":   "മായ്ക്കുക",
        "group_en":   "Backspace (leaf)",
        "type":       "Leaf / Control",
        "start_angle": 180,
        "sweep_angle":  30,
        "children":   [],
        "phonetic":   ["[DEL]"],
    },
    {
        "root_index": 11,
        "label":      "SPC",
        "group_ml":   "ഇടം",
        "group_en":   "Space (leaf)",
        "type":       "Leaf / Control",
        "start_angle": 210,
        "sweep_angle":  30,
        "children":   [],
        "phonetic":   ["[SPACE]"],
    },
    {
        "root_index": 12,
        "label":      "൦",
        "group_ml":   "അക്കങ്ങൾ",
        "group_en":   "Malayalam Numerals",
        "type":       "Drillable",
        "start_angle": 240,
        "sweep_angle":  30,
        "children":   ["൦","൧","൨","൩","൪","൫","൬","൭","൮","൯"],
        "phonetic":   ["0","1","2","3","4","5","6","7","8","9"],
    },
]

# ── Style helpers ─────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=11, color="000000", name="Calibri"):
    return Font(bold=bold, size=size, color=color, name=name)

def border(color=BORDER_CLR):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def write_cell(ws, row, col, value, bold=False, bg=None, fg="000000",
               size=11, h_align="center", wrap=False, border_on=True, num_fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = font(bold=bold, size=size, color=fg)
    cell.alignment = align(h=h_align, wrap=wrap)
    if bg:
        cell.fill = fill(bg)
    if border_on:
        cell.border = border()
    if num_fmt:
        cell.number_format = num_fmt
    return cell

# ── Sheet 1: Layout Overview ──────────────────────────────────────────────────
def build_overview(wb):
    ws = wb.create_sheet("Layout Overview")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # Title
    ws.merge_cells("A1:K1")
    t = ws["A1"]
    t.value     = "മലയാളം റേഡിയൽ കീബോർഡ് — Full Layout Overview"
    t.font      = Font(bold=True, size=16, color=HEADER_FG, name="Calibri")
    t.fill      = fill(TEAL_DARK)
    t.alignment = align(h="center")

    # Sub-header
    ws.merge_cells("A2:K2")
    s = ws["A2"]
    s.value     = "Edit the 'Character' and 'Phonetic' columns to rearrange — use family sheets for per-group editing"
    s.font      = Font(italic=True, size=10, color="546E7A", name="Calibri")
    s.fill      = fill(TEAL_LIGHT)
    s.alignment = align(h="center")

    # Column headers
    headers = [
        ("Seg #",           "A", 7),
        ("Root Label",      "B", 12),
        ("Group (ML)",      "C", 18),
        ("Group (EN)",      "D", 28),
        ("Type",            "E", 16),
        ("Start °",         "F", 9),
        ("Sweep °",         "G", 9),
        ("Char #",          "H", 8),
        ("Character",       "I", 12),
        ("Unicode",         "J", 12),
        ("Phonetic / Note", "K", 22),
    ]

    for i, (hdr, col_l, width) in enumerate(headers, start=1):
        write_cell(ws, 3, i, hdr, bold=True, bg=TEAL_MID, fg=HEADER_FG, size=10)
        set_col_width(ws, col_l, width)

    row = 4
    for seg in ROOT_SEGMENTS:
        bg      = FAMILY_COLORS.get(seg["label"], WHITE)
        n_chars = max(len(seg["children"]), 1)

        if seg["children"]:
            for ci, (char, phon) in enumerate(
                zip(seg["children"], seg["phonetic"]), start=1
            ):
                write_cell(ws, row, 1,  seg["root_index"],   bg=bg, h_align="center")
                write_cell(ws, row, 2,  seg["label"],         bg=bg, bold=True, size=14)
                write_cell(ws, row, 3,  seg["group_ml"],      bg=bg, h_align="left")
                write_cell(ws, row, 4,  seg["group_en"],      bg=bg, h_align="left")
                write_cell(ws, row, 5,  seg["type"],          bg=bg)
                write_cell(ws, row, 6,  seg["start_angle"],   bg=bg)
                write_cell(ws, row, 7,  seg["sweep_angle"],   bg=bg)
                write_cell(ws, row, 8,  ci,                   bg=CHAR_BG)
                write_cell(ws, row, 9,  char,                 bg=CHAR_BG, bold=True, size=14)
                write_cell(ws, row, 10, f"U+{ord(char):04X}", bg=CHAR_BG, fg="1565C0")
                write_cell(ws, row, 11, phon,                 bg=CHAR_BG, h_align="left")
                row += 1
        else:
            # Leaf (single row)
            char  = seg["label"] if seg["label"] not in ("⌫","SPC") else seg["label"]
            phon  = seg["phonetic"][0] if seg["phonetic"] else ""
            write_cell(ws, row, 1,  seg["root_index"],  bg=bg)
            write_cell(ws, row, 2,  seg["label"],        bg=bg, bold=True, size=13)
            write_cell(ws, row, 3,  seg["group_ml"],     bg=bg, h_align="left")
            write_cell(ws, row, 4,  seg["group_en"],     bg=bg, h_align="left")
            write_cell(ws, row, 5,  seg["type"],         bg=bg)
            write_cell(ws, row, 6,  seg["start_angle"],  bg=bg)
            write_cell(ws, row, 7,  seg["sweep_angle"],  bg=bg)
            write_cell(ws, row, 8,  "—",                 bg=bg)
            write_cell(ws, row, 9,  seg["label"],        bg=bg, bold=True, size=13)
            write_cell(ws, row, 10, "—",                 bg=bg)
            write_cell(ws, row, 11, phon,                bg=bg, h_align="left")
            row += 1

        # Spacer row between segments
        for c in range(1, 12):
            cell = ws.cell(row=row, column=c)
            cell.fill   = fill("ECEFF1")
            cell.border = border("CFD8DC")
        row += 1

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18

# ── Sheet 2: Root Ring ────────────────────────────────────────────────────────
def build_root_ring(wb):
    ws = wb.create_sheet("Root Ring")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value     = "Root Ring — 12 Segments (30° each)"
    t.font      = Font(bold=True, size=14, color=HEADER_FG, name="Calibri")
    t.fill      = fill(TEAL_DARK)
    t.alignment = align(h="center")

    ws.merge_cells("A2:H2")
    s = ws["A2"]
    s.value     = "Rearrange rows to change segment order on the wheel. Angles are auto-calculated (30° × position)."
    s.font      = Font(italic=True, size=10, color="546E7A")
    s.fill      = fill(TEAL_LIGHT)
    s.alignment = align(h="center")

    hdrs = ["Position","Label","Group (ML)","Group (EN)","Type","Start °","Sweep °","Children Count"]
    widths = [10, 12, 18, 30, 18, 10, 10, 16]
    for i, (h, w) in enumerate(zip(hdrs, widths), start=1):
        write_cell(ws, 3, i, h, bold=True, bg=TEAL_MID, fg=HEADER_FG, size=10)
        set_col_width(ws, get_column_letter(i), w)

    for r, seg in enumerate(ROOT_SEGMENTS, start=4):
        bg = FAMILY_COLORS.get(seg["label"], WHITE)
        write_cell(ws, r, 1, seg["root_index"],       bg=bg)
        write_cell(ws, r, 2, seg["label"],             bg=bg, bold=True, size=13)
        write_cell(ws, r, 3, seg["group_ml"],          bg=bg, h_align="left")
        write_cell(ws, r, 4, seg["group_en"],          bg=bg, h_align="left")
        write_cell(ws, r, 5, seg["type"],              bg=bg)
        write_cell(ws, r, 6, seg["start_angle"],       bg=bg)
        write_cell(ws, r, 7, seg["sweep_angle"],       bg=bg)
        write_cell(ws, r, 8, len(seg["children"]) or "leaf", bg=bg)

    ws.row_dimensions[1].height = 28

# ── Per-family sheets ─────────────────────────────────────────────────────────
def build_family_sheet(wb, seg):
    name = f"{seg['label']} {seg['group_en'].split('(')[0].strip()}"[:31]
    ws   = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    bg_header = FAMILY_COLORS.get(seg["label"], TEAL_LIGHT)

    # Title
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value     = f"{seg['group_ml']} ({seg['group_en']})"
    t.font      = Font(bold=True, size=14, color=TEAL_DARK, name="Calibri")
    t.fill      = fill(bg_header)
    t.alignment = align(h="center")

    # Metadata row
    ws.merge_cells("A2:F2")
    m = ws["A2"]
    m.value     = (f"Root Segment: {seg['label']}  |  "
                   f"Position #{seg['root_index']}  |  "
                   f"Angle: {seg['start_angle']}° – {seg['start_angle']+seg['sweep_angle']}°  |  "
                   f"{len(seg['children'])} characters in sub-ring")
    m.font      = Font(italic=True, size=10, color="546E7A")
    m.fill      = fill(bg_header)
    m.alignment = align(h="center")

    # Column headers
    hdrs   = ["Pos", "Character", "Unicode", "Phonetic / Transliteration",
              "Sub-angle °", "Notes / Rename"]
    widths = [6,      14,          12,         30,                           12,         30]
    for i, (h, w) in enumerate(zip(hdrs, widths), start=1):
        write_cell(ws, 3, i, h, bold=True, bg=TEAL_MID, fg=HEADER_FG, size=10)
        set_col_width(ws, get_column_letter(i), w)

    n = len(seg["children"])
    sub_sweep = round(360 / n, 2) if n else 0

    for ci, (char, phon) in enumerate(zip(seg["children"], seg["phonetic"]), start=1):
        row      = ci + 3
        sub_start = round(-90 + (ci-1) * sub_sweep, 2)
        bg_row   = CHAR_BG if ci % 2 == 0 else WHITE
        write_cell(ws, row, 1, ci,                    bg=bg_row)
        write_cell(ws, row, 2, char,                  bg=bg_row, bold=True, size=16)
        write_cell(ws, row, 3, f"U+{ord(char):04X}",  bg=bg_row, fg="1565C0")
        write_cell(ws, row, 4, phon,                  bg=bg_row, h_align="left")
        write_cell(ws, row, 5, f"{sub_start}°",       bg=bg_row, fg="757575")
        write_cell(ws, row, 6, "",                    bg=bg_row, h_align="left",
                   border_on=True)  # editable notes cell

    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 16

# ── Sheet: Instructions ───────────────────────────────────────────────────────
def build_instructions(wb):
    ws = wb.create_sheet("Instructions")
    ws.sheet_view.showGridLines = False
    set_col_width(ws, "A", 5)
    set_col_width(ws, "B", 55)
    set_col_width(ws, "C", 40)

    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value     = "How to Use This Workbook"
    t.font      = Font(bold=True, size=16, color=HEADER_FG)
    t.fill      = fill(TEAL_DARK)
    t.alignment = align(h="center")
    ws.row_dimensions[1].height = 30

    steps = [
        ("SHEET: Layout Overview",
         "Full flat list of every character in the keyboard. "
         "Edit the 'Character' column to swap letters. "
         "Edit 'Phonetic / Note' to annotate."),

        ("SHEET: Root Ring",
         "Shows all 12 root-level segments. Reorder rows here to change which "
         "family appears at which position (angle) on the wheel. "
         "Angles are 30° per segment starting from -90° (12 o'clock)."),

        ("SHEETS: Per-Family (ക, ച, ട …)",
         "Each drillable family has its own sheet. "
         "Edit the 'Character' or 'Phonetic' cells to rearrange sub-ring letters. "
         "Row order = clockwise order on the sub-ring wheel."),

        ("EDITING TIPS",
         "• Malayalam characters are Unicode — paste directly from any Unicode chart.\n"
         "• Sub-angle is auto-calculated: 360° ÷ number of children.\n"
         "• Leaf segments (⌫, SPC, ം) have no sub-ring.\n"
         "• After editing, share the file with the developer to update KeyboardLayout.kt."),

        ("UNICODE REFERENCE",
         "Malayalam Unicode block: U+0D00 – U+0D7F\n"
         "Vowels start at U+0D05 (അ)\n"
         "Consonants start at U+0D15 (ക)\n"
         "Chillu letters: U+0D7A – U+0D7F\n"
         "Numerals: U+0D66 (൦) – U+0D6F (൯)"),
    ]

    row = 3
    for title, body in steps:
        ws.merge_cells(f"B{row}:C{row}")
        h = ws.cell(row=row, column=2, value=title)
        h.font      = Font(bold=True, size=11, color=HEADER_FG)
        h.fill      = fill(TEAL_MID)
        h.alignment = align(h="left", wrap=True)
        ws.row_dimensions[row].height = 20
        row += 1

        ws.merge_cells(f"B{row}:C{row}")
        b = ws.cell(row=row, column=2, value=body)
        b.font      = Font(size=10, color="212121")
        b.fill      = fill(GREY_LIGHT)
        b.alignment = align(h="left", wrap=True)
        ws.row_dimensions[row].height = max(15 * body.count("\n") + 18, 30)
        row += 2

# ── Main ──────────────────────────────────────────────────────────────────────
OUTPUT = "/Users/jaisonjacob89/Desktop/RadialKeyboardApp/Malayalam_Radial_Keyboard_Layout.xlsx"

wb = Workbook()
wb.remove(wb.active)  # remove default sheet

build_overview(wb)
build_root_ring(wb)

for seg in ROOT_SEGMENTS:
    if seg["children"]:
        build_family_sheet(wb, seg)

build_instructions(wb)

# Set Layout Overview as the active sheet
wb.active = wb["Layout Overview"]

wb.save(OUTPUT)
print(f"✓ Saved: {OUTPUT}")
print(f"  Sheets: {[s.title for s in wb.worksheets]}")
